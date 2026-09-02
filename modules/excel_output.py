#!/usr/bin/env python3

import os

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)

from openpyxl.utils import (
    get_column_letter
)


EXCEL_MAX_ROWS = 1_048_576

MAX_DATA_ROWS = EXCEL_MAX_ROWS - 1


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)


# ============================================================
# HEADER
# ============================================================

def style_header(ws):

    for cell in ws[1]:

        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[1].height = 30


# ============================================================
# WIDTH
# ============================================================

def set_widths(ws):

    for column_cells in ws.columns:

        if not column_cells:
            continue

        try:

            column_index = (
                column_cells[0].column
            )

            column_letter = (
                get_column_letter(
                    column_index
                )
            )

            max_length = 0

            for cell in column_cells[:200]:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(
                            str(cell.value)
                        )
                    )

            width = min(
                max(
                    max_length + 2,
                    14
                ),
                70
            )

            ws.column_dimensions[
                column_letter
            ].width = width

        except Exception:

            pass


# ============================================================
# TABLE WRITER
# ============================================================

def write_table_split(
    workbook,
    sheet_name,
    headers,
    rows
):

    sheet_names = []

    sheet_number = 1

    ws = None

    data_rows = 0

    if not rows:

        ws = workbook.create_sheet(
            sheet_name
        )

        ws.append(
            headers
        )

        style_header(ws)

        ws.freeze_panes = "A2"

        sheet_names.append(
            sheet_name
        )

        return sheet_names

    for row in rows:

        if ws is None:

            if sheet_number == 1:

                current_name = sheet_name

            else:

                current_name = (
                    f"{sheet_name}_{sheet_number}"
                )

            ws = workbook.create_sheet(
                current_name
            )

            ws.append(
                headers
            )

            style_header(ws)

            ws.freeze_panes = "A2"

            sheet_names.append(
                current_name
            )

            data_rows = 0

        if data_rows >= MAX_DATA_ROWS:

            sheet_number += 1

            current_name = (
                f"{sheet_name}_{sheet_number}"
            )

            ws = workbook.create_sheet(
                current_name
            )

            ws.append(
                headers
            )

            style_header(ws)

            ws.freeze_panes = "A2"

            sheet_names.append(
                current_name
            )

            data_rows = 0

        ws.append(
            list(row)
        )

        data_rows += 1

    for name in sheet_names:

        ws = workbook[name]

        ws.auto_filter.ref = (
            ws.dimensions
        )

        set_widths(ws)

    return sheet_names


# ============================================================
# CREATE EXCEL
# ============================================================

def create_excel(

    output_file,

    protein_go,

    annotation_rows,

    slim_rows,

    interpro_ids,

    go_terms,

    blast_rows=None,

    interpro_summary_rows=None,

    protein_slim=None

):

    if blast_rows is None:
        blast_rows = []

    if interpro_summary_rows is None:
        interpro_summary_rows = []

    if protein_slim is None:
        protein_slim = {}

    print()
    print("=" * 70)
    print("CREATING FINAL EXCEL WORKBOOK")
    print("=" * 70)


    wb = Workbook()

    default_sheet = wb.active

    wb.remove(
        default_sheet
    )


    # ========================================================
    # SUMMARY
    # ========================================================

    summary = wb.create_sheet(
        "Summary"
    )

    summary.append([
        "Protein Annotation Pipeline"
    ])

    summary.append([])

    summary.append([
        "Organism",
        "Fusarium"
    ])

    summary.append([
        "Unique proteins",
        len(protein_go)
    ])

    summary.append([
        "InterPro proteins",
        len([
            p for p, v in interpro_ids.items()
            if v
        ])
    ])

    summary.append([
        "InterPro IDs",
        len(
            set(
                ipr
                for values in interpro_ids.values()
                for ipr in values
            )
        )
    ])

    summary.append([
        "BLAST best hits",
        len(blast_rows)
    ])

    summary.append([
        "GO annotation rows",
        len(annotation_rows)
    ])

    summary.append([
        "GO-Slim rows",
        len(slim_rows)
    ])

    summary.append([])

    summary.append([
        "BLAST database",
        "UniProtKB/Swiss-Prot"
    ])

    summary.append([
        "InterProScan",
        "InterProScan 6"
    ])

    summary.append([
        "Excel row-limit protection",
        "Enabled"
    ])

    summary["A1"].fill = HEADER_FILL

    summary["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )

    summary.column_dimensions[
        "A"
    ].width = 38

    summary.column_dimensions[
        "B"
    ].width = 70


    # ========================================================
    # INTERPRO SUMMARY
    # ========================================================

    interpro_headers = [

        "Protein",

        "InterPro_ID",

        "InterPro_Description",

        "Pfam",

        "GO_Terms",

        "Pathways",

        "Analysis"

    ]

    interpro_sheets = write_table_split(

        wb,

        "InterPro_Summary",

        interpro_headers,

        interpro_summary_rows

    )

    print(
        "InterPro summary sheets:",
        ", ".join(interpro_sheets)
    )


    # ========================================================
    # BLAST
    # ========================================================

    blast_headers = [

        "Protein",

        "SwissProt_Accession",

        "Description",

        "Percent_Identity",

        "Alignment_Length",

        "Query_Length",

        "Query_Start",

        "Query_End",

        "Evalue",

        "Bitscore",

        "Query_Coverage",

        "Classification"

    ]

    blast_sheets = write_table_split(

        wb,

        "BLAST_SwissProt",

        blast_headers,

        blast_rows

    )

    print(
        "BLAST sheets:",
        ", ".join(blast_sheets)
    )


    # ========================================================
    # PROTEIN GO
    # ========================================================

    protein_headers = [

        "Protein",

        "InterPro_IDs",

        "GO_Biological_Process",

        "GO_Molecular_Function",

        "GO_Cellular_Component",

        "GO_BP_Names",

        "GO_MF_Names",

        "GO_CC_Names",

        "GO_Slim_Biological_Process",

        "GO_Slim_Molecular_Function",

        "GO_Slim_Cellular_Component",

        "GO_Slim_BP_Names",

        "GO_Slim_MF_Names",

        "GO_Slim_CC_Names"

    ]

    protein_rows = []

    for protein in sorted(
        protein_go
    ):

        categories = protein_go[
            protein
        ]

        bp = categories.get(
            "Biological Process",
            set()
        )

        mf = categories.get(
            "Molecular Function",
            set()
        )

        cc = categories.get(
            "Cellular Component",
            set()
        )

        slim = protein_slim.get(
            protein,
            {}
        )

        sbp = slim.get(
            "Biological Process",
            set()
        )

        smf = slim.get(
            "Molecular Function",
            set()
        )

        scc = slim.get(
            "Cellular Component",
            set()
        )

        def names(go_ids):

            values = []

            for go_id in sorted(go_ids):

                name = go_terms.get(
                    go_id,
                    {}
                ).get(
                    "name",
                    ""
                )

                if name:
                    values.append(
                        name
                    )

            return "; ".join(
                values
            ) or "-"

        protein_rows.append([

            protein,

            "; ".join(
                sorted(
                    interpro_ids.get(
                        protein,
                        set()
                    )
                )
            ) or "-",

            "; ".join(
                sorted(bp)
            ) or "-",

            "; ".join(
                sorted(mf)
            ) or "-",

            "; ".join(
                sorted(cc)
            ) or "-",

            names(bp),

            names(mf),

            names(cc),

            "; ".join(
                sorted(sbp)
            ) or "-",

            "; ".join(
                sorted(smf)
            ) or "-",

            "; ".join(
                sorted(scc)
            ) or "-",

            names(sbp),

            names(smf),

            names(scc)

        ])


    protein_sheets = write_table_split(

        wb,

        "Protein_GO",

        protein_headers,

        protein_rows

    )

    print(
        "Protein GO sheets:",
        ", ".join(protein_sheets)
    )


    # ========================================================
    # GO ANNOTATIONS
    # ========================================================

    go_headers = [

        "Protein",

        "InterPro_ID",

        "GO_ID",

        "GO_Name",

        "GO_Category",

        "Source",

        "Evidence"

    ]

    go_sheets = write_table_split(

        wb,

        "GO_Annotations",

        go_headers,

        annotation_rows

    )

    print(
        "GO annotation sheets:",
        ", ".join(go_sheets)
    )


    # ========================================================
    # GO-SLIM
    # ========================================================

    goslim_headers = [

        "Protein",

        "GO_ID",

        "GO_Slim_ID",

        "GO_Slim_Name",

        "GO_Slim_Category"

    ]

    goslim_sheets = write_table_split(

        wb,

        "GO_Slim",

        goslim_headers,

        slim_rows

    )

    print(
        "GO-Slim sheets:",
        ", ".join(goslim_sheets)
    )


    # ========================================================
    # SAVE
    # ========================================================

    output_directory = os.path.dirname(
        output_file
    )

    if output_directory:

        os.makedirs(
            output_directory,
            exist_ok=True
        )

    print()
    print("Saving Excel:")
    print(output_file)

    wb.save(
        output_file
    )

    wb.close()


    size_mb = (

        os.path.getsize(
            output_file
        )
        /
        (1024 ** 2)

    )


    print()
    print("=" * 70)
    print("EXCEL CREATION COMPLETED")
    print("=" * 70)

    print(
        f"Unique proteins : {len(protein_go):,}"
    )

    print(
        f"InterPro rows   : {len(interpro_summary_rows):,}"
    )

    print(
        f"BLAST hits      : {len(blast_rows):,}"
    )

    print(
        f"GO rows         : {len(annotation_rows):,}"
    )

    print(
        f"GO-Slim rows    : {len(slim_rows):,}"
    )

    print(
        f"Excel size      : {size_mb:.1f} MB"
    )

    print("=" * 70)
